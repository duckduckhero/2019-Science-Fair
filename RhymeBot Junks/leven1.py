import g2p
import openpyxl

from openpyxl import Workbook

wb = openpyxl.load_workbook('dataset.xlsx')
ws = wb.active
write_wb = Workbook()

write_ws = write_wb.active

def leven(aText, bText):
    aLen = len(aText)+1
    bLen = len(bText)+1

    array = [ [] for a in range(aLen) ]
    for i in range(aLen):
        array[i] = [0 for a in range(bLen)] #배열 초기화
    for i in range(bLen): #공집합까지의 거리로 채움
        array[0][i] = i
    for i in range(aLen): #공집합까지의 거리로 채움
        array[i][0] = i
    cost = 0
    for i in range(1, aLen):
        for j in range(1, bLen):
            if aText[i-1] != bText[j-1]:
                cost=1
            else :
                cost=0
            addNum = array[i-1][j] + 1
            minusNum = array[i][j-1] + 1
            modiNum = array[i-1][j-1] + cost
            minNum = min([addNum, minusNum, modiNum])
            array[i][j] = minNum

    for i in range(aLen):
        for j in range(bLen):
            print(array[i][j], end='')
        print()
    return array[aLen-1][bLen-1]

cnt=1
for r in ws.rows:
    if r[0].value!='단어1':
        w1 = r[0].value
        w2 = r[1].value
        print(w1)
        print(w2)
        w1 = g2p.runKoG2P(w1, 'rulebook.txt')
        w2 = g2p.runKoG2P(w2, 'rulebook.txt')
        w1 = w1.split()
        w2 = w2.split()
        print('입력된 단어들의 음운 구성')
        print(w1)
        print(w2)
        res = leven(w1, w2)
        write_ws.cell(cnt, 1, r[2].value)
        write_ws.cell(cnt, 2, res)
        cnt=cnt+1

write_wb.save('leven1_res.xlsx')
