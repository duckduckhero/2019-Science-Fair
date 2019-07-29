import g2p
#자음과 모음에서 가장 큰 Subsequence 길이 찾은 다음에 둘이 더한다
import openpyxl

from openpyxl import Workbook

wb = openpyxl.load_workbook('검증셋.xlsx')
ws = wb.active
#write_wb = Workbook()

#write_ws = write_wb.active

vowels = ['ii','ee','qq','aa','xx','vv','uu','oo','ye','yq','ya','yv','yu','yo','wi','wo','wq','we','wa','wv','xi']
consonants = ['p0','ph','pp','t0','th','tt','k0','kh','kk','s0,','ss','h0','c0','mm','nn','rr','pf','ph','tf','th','kf','kh','s0','ss','h0','c0','mf','nf','ng','ll','ks','nc','nh','lk','lm','lt','lp','lh','ps']

v1 = ['ee','qq', 'ye', 'yq', 'wq','wo', 'we', 'xi'] #ㅔ 계열 모음
v2 = ['ii','wi'] #ㅣ 계열 모음
v3 = ['aa', 'ya', 'wa'] # ㅏ 계열 모음
v4 = ['uu', 'yu'] # ㅜ 계열 모음
v5 = ['vv', 'oo', 'yv', 'wv'] #ㅓ계열 모음

c1 = ['p0', 'ph', 'pp'] # ㅂ 계열 자음
c2 = ['t0', 'th', 'tt'] # ㄷ 계열 자음
c3 = ['k0', 'kh', 'kk'] # ㄱ 계열 자음
c4 = ['s0', 'ss'] # ㅅ 계열 자음
c5 = ['c0', 'ch', 'cc'] # ㅈ 계열 자음

def vowelconvt(vow1): #모음을 해당 그룹의 대표 모음으로 치환하는 함수
    length = len(vow1)
    for i in range(length):
        if v1.count(vow1[i])==1: vow1[i]='ee'
        if v2.count(vow1[i])==1: vow1[i]='ii'
        if v3.count(vow1[i])==1: vow1[i]='aa'
        if v4.count(vow1[i])==1: vow1[i]='uu'
        if v5.count(vow1[i])==1: vow1[i]='vv'
    return vow1

def conconvt(con1):
    length = len(con1)
    for i in range(length):
        if c1.count(con1[i])==1: con1[i]='p0'
        if c2.count(con1[i])==1: con1[i]='t0'
        if c3.count(con1[i])==1: con1[i]='k0'
        if c4.count(con1[i])==1: con1[i]='s0'
        if c5.count(con1[i])==1: con1[i]='c0'
    return con1

def vowel(word1): #단어 음절에서 모음만 빼서 모으는 함
    length = len(word1)
    w1v = []
    for i in range(length):
        if vowels.count(word1[i])!=0:
            w1v.append(word1[i])
    return w1v

def consonant(word1): #단어 음절에서 자음만 빼서 모으는 함수
    length = len(word1)
    w1c = []
    for i in range(length):
        if consonants.count(word1[i])!=0:
            w1c.append(word1[i])
    return w1c

def lcs_length(a, b):
    table = [[0] * (len(b) + 1) for _ in range(len(a) + 1)]
    for i, ca in enumerate(a, 1):
        for j, cb in enumerate(b, 1):
            table[i][j] = (
                table[i - 1][j - 1] + 1 if ca == cb else
                max(table[i][j - 1], table[i - 1][j]))
    return table[-1][-1]

cnt=0
flag=0
for r in ws.rows:
    if r[0].value!='단어1':
        w1 = r[0].value
        w2 = r[1].value
        w1 = g2p.runKoG2P(w1, 'rulebook.txt')
        w2 = g2p.runKoG2P(w2, 'rulebook.txt')
        w1 = w1.split()
        w2 = w2.split()
        print('변환하기 전의 음운 구성')
        print(w1)
        print(w2)
        w1 = vowelconvt(w1)
        w2 = vowelconvt(w2)
        w1 = conconvt(w1)
        w2 = conconvt(w2)
        print('변환 후의 음운 구성')
        print(w1)
        print(w2)
        w1v = vowel(w1)
        w1c = consonant(w1)
        w2v = vowel(w2)
        w2c = consonant(w2)
        print('자음 추출')
        print(w1c)
        print(w2c)
        print('모음 추출')
        print(w1v)
        print(w2v)
        vlcs = lcs_length(w1v, w2v) #모음 최장 연속 구간
        clcs = lcs_length(w1c, w2c) #자음 최장 연속 구간
        print("모음 최장 겹침 : "+ str(vlcs))
        print("자음 최장 겹침 : "+ str(clcs))
        print("둘의 합계 : "+str(vlcs+clcs))
        res = (vlcs+clcs)/max(len(w1), len(w2))
        #write_ws.cell(cnt, 1, r[2].value)
        #write_ws.cell(cnt, 2, res)
        #cnt=cnt+1
        if res>=0.62:
            flag=1
        if flag==r[2].value:
            cnt=cnt+1
        flag=0

print("총 맞힌 횟수 : "+ str(cnt))
print("정확률 : "+str(cnt/100))

'''
flag=0
w1 = input()
w2 = input()
w1 = g2p.runKoG2P(w1, 'rulebook.txt')
w2 = g2p.runKoG2P(w2, 'rulebook.txt')
w1 = w1.split()
w2 = w2.split()
print('변환하기 전의 음운 구성')
print(w1)
print(w2)
w1 = vowelconvt(w1)
w2 = vowelconvt(w2)
w1 = conconvt(w1)
w2 = conconvt(w2)
print('변환 후의 음운 구성')
print(w1)
print(w2)
w1v = vowel(w1)
w1c = consonant(w1)
w2v = vowel(w2)
w2c = consonant(w2)
print('자음 추출')
print(w1c)
print(w2c)
print('모음 추출')
print(w1v)
print(w2v)
vlcs = lcs_length(w1v, w2v) #모음 최장 연속 구간
clcs = lcs_length(w1c, w2c) #자음 최장 연속 구간
print("모음 최장 겹침 : "+ str(vlcs))
print("자음 최장 겹침 : "+ str(clcs))
print("둘의 합계 : "+str(vlcs+clcs))
res = (vlcs+clcs)/max(len(w1), len(w2))
if res>=0.9:
    flag=1
if flag==1:
    print("Rhymes")
else :
    print("Does Not Rhyme")
'''
#write_wb.save('lcs2_res.xlsx')
